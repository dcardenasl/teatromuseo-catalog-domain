import json
import os
import openpyxl

def excel_to_json():
    xlsx_path = "/Users/davidcardenas/Developer/PHP/teatromuseo/docs/2026-05-19-plantilla-fichas-coleccion (1).xlsx"
    if not os.path.exists(xlsx_path):
        print(f"Error: Excel file not found at {xlsx_path}")
        return False
        
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "Fichas de Colección" not in wb.sheetnames:
        print("Error: Sheet 'Fichas de Colección' not found in workbook")
        return False
        
    ws = wb["Fichas de Colección"]
    
    # Read technical headers from row 1
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    headers = [h.strip() if h else None for h in headers]
    
    records = []
    # Read data starting from row 3
    for row_idx, row in enumerate(ws.iter_rows(min_row=3), start=3):
        row_values = [cell.value for cell in row]
        if not any(row_values):
            continue # skip empty rows
            
        record = {}
        for col_idx, val in enumerate(row_values):
            if col_idx < len(headers) and headers[col_idx]:
                key = headers[col_idx]
                # Convert Excel cell values to clean types
                if val is None:
                    record[key] = ""
                else:
                    record[key] = str(val).strip()
                    
        # Verify if record has at least name
        if record.get("nombre"):
            records.append(record)
            
    # Output path
    output_dir = "/Users/davidcardenas/Developer/PHP/teatromuseo/teatromuseo-catalog-domain/writable"
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, "temp_excel_data.json")
    
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=4)
        
    print(f"Successfully converted {len(records)} records to {output_path}")
    return True

if __name__ == "__main__":
    excel_to_json()
